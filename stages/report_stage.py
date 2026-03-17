# file: stages/report_stage.py
"""
REPORT stage
report.source 설정에 따라 소스 DB 또는 target DB에서 SQL 실행 → CSV 저장 → Excel 변환.
skip_sql: true 시 DB 연결 없이 기존 CSV를 바로 union → Excel 처리.

job.yml 설정:
  report:
    source: target          # target(기본) / oracle / vertica
    skip_sql: false         # true: DB 연결 없이 csv_union_dir의 CSV만 사용
    csv_union_dir: data/export  # skip_sql=true 시 union할 CSV 소스 폴더
    export_csv:
      enabled: true
      sql_dir: sql/report/
      out_dir: data/report/
      compression: none     # none(기본) / gzip
    excel:
      enabled: true
      out_dir: data/report/
      max_files: 10
      sheet_mode: merge        # merge(기본): 동일 테이블 → 같은 시트에 union
                               # separate: CSV 파일별 개별 시트 생성
      csv_filter:              # CSV 파일명 LIKE 필터 (대소문자 무시)
                               # 문자열: "contract" → 파일명에 포함된 것만
                               # 리스트: ["contract", "order"] → OR 조건
"""

import csv
import gzip
import re
import time
from datetime import datetime
from pathlib import Path

from engine.connection import connect_target, set_session_schema
from engine.context import RunContext
from engine.path_utils import resolve_path
from engine.sql_utils import sort_sql_files, render_sql, strip_sql_prefix, read_sql_file
from stages.task_tracking import (
    make_task_key, init_run_info, update_task_status, load_failed_tasks,
)


def run(ctx: RunContext):
    logger = ctx.logger
    # logger.info("REPORT stage start")

    report_cfg = ctx.job_config.get("report")
    if not report_cfg:
        logger.info("REPORT stage skipped (no config)")
        return

    if ctx.mode == "plan":
        logger.info("REPORT stage skipped (plan mode)")
        return

    # 스테이지별 독립 params / param_mode
    stage_params = ctx.get_stage_params("report")
    stage_param_mode = ctx.get_stage_param_mode("report")

    # ── run_info.json 초기화 & retry ────────────────────────
    tracking_base = resolve_path(ctx, report_cfg.get("tracking_dir", ctx.get_default("tracking_dir_report")))
    run_info_dir = tracking_base / ctx.run_id
    run_info_path = run_info_dir / "run_info.json"

    init_run_info(run_info_path, job_name=ctx.job_name, run_id=ctx.run_id,
                  stage="report", mode=ctx.mode, params=ctx.params)

    failed_task_keys = None
    if ctx.mode == "retry":
        failed_task_keys = load_failed_tasks(
            tracking_base, ctx.run_id, stage="report")

    generated_csvs = []
    report_success = 0
    report_failed = 0
    report_skipped = 0

    skip_sql = bool(report_cfg.get("skip_sql", False))

    if skip_sql:
        # DB 연결 없이 csv_union_dir 의 CSV 파일들을 바로 사용
        # 디폴트: export.out_dir (export 직후 바로 union하는 패턴)
        csv_union_dir = report_cfg.get("csv_union_dir")
        if not csv_union_dir:
            export_cfg = ctx.job_config.get("export", {})
            export_out = export_cfg.get("out_dir", ctx.get_default("export_out_dir"))
            csv_union_dir = str(Path(export_out))
        union_dir = resolve_path(ctx, csv_union_dir)
        if union_dir.exists():
            generated_csvs = sorted(
                p for p in union_dir.iterdir()
                if p.is_file() and (p.name.endswith(".csv") or p.name.endswith(".csv.gz"))
            )
            logger.info("REPORT skip_sql=true | csv_union_dir=%s files=%d",
                        union_dir, len(generated_csvs))
        else:
            logger.warning("REPORT skip_sql=true | csv_union_dir not found: %s", union_dir)
    else:
        export_csv_cfg = report_cfg.get("export_csv", {})
        if export_csv_cfg.get("enabled", False):
            generated_csvs, csv_failed, csv_skipped = _run_csv_export(
                ctx, report_cfg, export_csv_cfg,
                stage_params=stage_params,
                stage_param_mode=stage_param_mode,
                run_info_path=run_info_path,
                failed_task_keys=failed_task_keys,
            )
            report_success += len(generated_csvs)
            report_failed += csv_failed
            report_skipped += csv_skipped
        else:
            logger.info("REPORT csv export skipped")

    excel_cfg = report_cfg.get("excel", {})
    if excel_cfg.get("enabled", False):
        try:
            _run_excel_export(ctx, report_cfg, excel_cfg, generated_csvs)
            report_success += 1
        except Exception as e:
            logger.error("REPORT excel FAILED: %s", e)
            report_failed += 1
    else:
        logger.info("REPORT excel export skipped")

    ctx.report_stage_result("report", success=report_success, failed=report_failed,
                            skipped=report_skipped)


# ────────────────────────────────────────────────────────────
# CSV Export
# ────────────────────────────────────────────────────────────

def _run_csv_export(ctx, report_cfg, cfg, *,
                    stage_params=None, stage_param_mode="product",
                    run_info_path=None, failed_task_keys=None) -> tuple:
    """CSV export. 반환: (generated_files, failed_count, skipped_count)."""
    from stages.export_stage import expand_params
    from engine.sql_utils import detect_used_params

    logger = ctx.logger
    if stage_params is None:
        stage_params = ctx.get_stage_params("report")
    sql_dir = resolve_path(ctx, cfg.get("sql_dir", ctx.get_default("report_sql_dir")))
    out_dir  = resolve_path(ctx, cfg.get("out_dir",  ctx.get_default("report_out_dir")))
    compression = (cfg.get("compression") or "none").strip().lower()

    out_dir.mkdir(parents=True, exist_ok=True)

    if not sql_dir.exists():
        logger.warning("REPORT sql_dir not found: %s", sql_dir)
        return [], 0, 0

    sql_files = sort_sql_files(sql_dir)
    if not sql_files:
        logger.info("REPORT no SQL files in %s", sql_dir)
        return [], 0, 0

    # ── --include-report 필터 적용 ──────────────────────────
    include_patterns = getattr(ctx, "include_report_patterns", []) or []
    if include_patterns:
        def _matches(f):
            rel  = f.relative_to(sql_dir).as_posix().lower()
            stem = f.stem.lower()
            return any(
                pat.lower() in rel or pat.lower() in stem
                for pat in include_patterns
            )
        before = len(sql_files)
        sql_files = [f for f in sql_files if _matches(f)]
        logger.info(
            "REPORT --include filter applied: %d -> %d files (patterns: %s)",
            before, len(sql_files), include_patterns
        )
        if not sql_files:
            logger.warning("REPORT --include filter resulted in no SQL files (patterns=%s)", include_patterns)
            return [], 0, 0

    # report.source 결정: 기본값은 "target"
    report_source = (report_cfg.get("source") or "target").strip().lower()
    conn, conn_type, conn_label = _open_connection(ctx, report_source)

    # 세션 스키마 설정: report.schema 우선, 없으면 target.schema fallback
    schema = (report_cfg.get("schema") or "").strip() \
             or (ctx.job_config.get("target", {}).get("schema") or "").strip() \
             or ""
    if schema:
        set_session_schema(conn, conn_type, schema, logger)

    logger.info(
        "REPORT csv export | db=%s sql_dir=%s out_dir=%s files=%d compression=%s",
        conn_label, sql_dir, out_dir, len(sql_files), compression,
    )

    generated = []
    csv_failed = 0
    csv_skipped = 0
    total = len(sql_files)
    ext = ".csv.gz" if compression == "gzip" else ".csv"

    try:
        for i, sql_file in enumerate(sql_files, 1):
            sql_text = read_sql_file(sql_file)

            # SQL별 사용 파라미터만 확장
            used_keys = detect_used_params(sql_text, stage_params)
            relevant_params = {k: v for k, v in stage_params.items() if k in used_keys}
            param_sets = expand_params(relevant_params, mode=stage_param_mode) if relevant_params else [{}]

            for pi, param_set in enumerate(param_sets, 1):
                task_key = make_task_key(sql_file, param_set)

                # retry 모드: 이전에 성공한 task는 건너뛰기
                if failed_task_keys is not None and task_key not in failed_task_keys:
                    logger.info("REPORT [%d/%d][%d/%d] %s RETRY skip (succeeded)",
                                i, total, pi, len(param_sets), sql_file.name)
                    csv_skipped += 1
                    continue

                full_params = {**stage_params, **param_set}
                rendered = render_sql(sql_text, full_params)

                # 파라미터가 여러 세트면 파일명에 파라미터값 포함
                if len(param_sets) > 1:
                    suffix = "_".join(str(v) for v in param_set.values())
                    out_file = out_dir / f"{sql_file.stem}_{suffix}{ext}"
                else:
                    out_file = out_dir / (sql_file.stem + ext)

                label = f"REPORT [{i}/{total}]" if len(param_sets) == 1 \
                    else f"REPORT [{i}/{total}][{pi}/{len(param_sets)}]"

                if run_info_path:
                    update_task_status(run_info_path, task_key, "running")

                logger.info("%s %s → %s", label, sql_file.name, out_file.name)
                start = time.time()
                try:
                    rows = _export_to_csv(conn, conn_type, rendered, out_file, compression)
                    elapsed = time.time() - start
                    logger.info("%s done | rows=%d elapsed=%.2fs", label, rows, elapsed)
                    generated.append(out_file)
                    if run_info_path:
                        update_task_status(run_info_path, task_key, "success",
                                           rows=rows, elapsed=elapsed)
                except Exception as e:
                    elapsed = time.time() - start
                    logger.error("%s FAILED (%.2fs): %s", label, elapsed, e)
                    csv_failed += 1
                    if run_info_path:
                        update_task_status(run_info_path, task_key, "failed", error=str(e))
    finally:
        conn.close()

    logger.info("REPORT csv summary | success=%d failed=%d skipped=%d",
                len(generated), csv_failed, csv_skipped)
    return generated, csv_failed, csv_skipped


def _open_connection(ctx, report_source: str):
    """
    report_source 에 따라 연결 반환.
      "target"          → job의 target DB (DuckDB/SQLite3/Oracle MYDATA)
      "oracle"/"vertica" → 소스 DB
    반환: (conn, conn_type, label)   실패 시 (None, None, None)
    """
    logger = ctx.logger

    if report_source == "target":
        target_cfg = ctx.job_config.get("target", {})
        return connect_target(ctx, target_cfg)

    # source DB (oracle / vertica)
    from adapters.sources.oracle_client import init_oracle_client, get_oracle_conn
    from adapters.sources.vertica_client import get_vertica_conn

    source_sel = ctx.job_config.get("source", {})
    src_type  = source_sel.get("type", "oracle")
    host_name = source_sel.get("host", "")
    env_cfg   = ctx.env_config

    if src_type == "oracle":
        oracle_cfg = env_cfg["sources"]["oracle"]
        host_cfg = oracle_cfg["hosts"].get(host_name)
        if not host_cfg:
            raise RuntimeError(f"Oracle host not found: {host_name}")
        init_oracle_client(oracle_cfg)
        dsn = host_cfg.get("dsn") or f"{host_cfg.get('host')}:{host_cfg.get('port', 1521)}/{host_cfg.get('service_name', '')}"
        label = f"oracle source ({dsn})"
        return get_oracle_conn(host_cfg), "oracle", label

    elif src_type == "vertica":
        vertica_cfg = env_cfg["sources"]["vertica"]
        host_cfg = vertica_cfg["hosts"].get(host_name)
        if not host_cfg:
            raise RuntimeError(f"Vertica host not found: {host_name}")
        label = f"vertica ({host_name})"
        return get_vertica_conn(host_cfg), "vertica", label

    else:
        raise ValueError(f"REPORT: unsupported source type: {src_type}")


def _export_to_csv(conn, conn_type: str, sql_text: str, out_file: Path, compression: str) -> int:
    """SQL 실행 결과를 CSV 저장. row 수 반환."""
    open_fn = gzip.open if compression == "gzip" else open
    row_count = 0

    if conn_type == "duckdb":
        rel = conn.execute(sql_text)
        columns = [d[0] for d in rel.description]
        with open_fn(out_file, "wt", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            while True:
                batch = rel.fetchmany(10000)
                if not batch:
                    break
                for row in batch:
                    writer.writerow(["" if v is None else str(v) for v in row])
                    row_count += 1
    else:
        cur = conn.cursor()
        try:
            if conn_type == "oracle":
                cur.arraysize = 10000
            cur.execute(sql_text)
            columns = [d[0] for d in cur.description]
            with open_fn(out_file, "wt", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(columns)
                while True:
                    batch = cur.fetchmany(10000)
                    if not batch:
                        break
                    for row in batch:
                        writer.writerow(["" if v is None else str(v) for v in row])
                        row_count += 1
        finally:
            cur.close()

    return row_count


# ────────────────────────────────────────────────────────────
# Excel Export (pandas 기반, csv_to_excel 로직)
# ────────────────────────────────────────────────────────────

def _get_excel_output_path(ctx, out_dir: Path, job_name: str, max_files: int) -> Path:
    logger = ctx.logger
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%y%m%d")
    base = f"{job_name}_{ts}"
    pattern = re.compile(rf"{re.escape(base)}_(\d+)\.xlsx")

    existing = []
    for f in out_dir.glob(f"{base}_*.xlsx"):
        m = pattern.fullmatch(f.name)
        if m:
            existing.append((int(m.group(1)), f))

    next_idx = max((i for i, _ in existing), default=0) + 1
    out = out_dir / f"{base}_{next_idx}.xlsx"

    existing_sorted = sorted(existing, key=lambda x: x[1].stat().st_mtime)
    while len(existing_sorted) >= max_files:
        _, old = existing_sorted.pop(0)
        old.unlink()
        logger.info("REPORT excel: deleting old file %s", old.name)

    return out


def _run_excel_export(ctx, report_cfg, cfg, csv_files: list):
    logger = ctx.logger
    out_dir_str = cfg.get("out_dir") or report_cfg.get("export_csv", {}).get("out_dir", ctx.get_default("report_out_dir"))
    out_dir = resolve_path(ctx, out_dir_str)
    max_files = int(cfg.get("max_files", 10))

    # ── csv_filter: LIKE 필터링 (대소문자 무시) ──
    csv_filter = cfg.get("csv_filter")
    if csv_filter:
        keywords = csv_filter if isinstance(csv_filter, list) else [csv_filter]
        keywords = [str(k).lower() for k in keywords]
        before = len(csv_files)
        csv_files = [f for f in csv_files
                     if any(k in f.name.lower() for k in keywords)]
        logger.info("REPORT excel csv_filter=%s | %d → %d files",
                    keywords, before, len(csv_files))

    if not csv_files:
        logger.warning("REPORT excel: no target CSV found, skip")
        return

    output_path = _get_excel_output_path(ctx, out_dir, ctx.job_name, max_files)
    logger.info("REPORT excel export | output=%s files=%d", output_path.name, len(csv_files))
    start = time.time()

    try:
        import pandas as pd
        from pandas.api.types import is_integer_dtype, is_float_dtype
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        # ── sheet_mode 에 따라 시트 구성 ──
        sheet_mode = (cfg.get("sheet_mode") or "merge").strip().lower()
        from collections import OrderedDict
        sheet_frames: OrderedDict[str, list[pd.DataFrame]] = OrderedDict()

        for csv_file in csv_files:
            raw_stem = csv_file.stem.replace(".csv", "")
            base_name = raw_stem.split("__", 1)[0]

            if sheet_mode == "separate":
                # 파일별 개별 시트: 전체 stem 사용 (31자 제한)
                sheet_name = strip_sql_prefix(raw_stem).upper()[:31]
                # 동일 이름 충돌 방지
                if sheet_name in sheet_frames:
                    idx = 2
                    while f"{sheet_name[:28]}_{idx}" in sheet_frames:
                        idx += 1
                    sheet_name = f"{sheet_name[:28]}_{idx}"
            else:
                # merge(기본): 동일 테이블명 → 같은 시트에 union
                sheet_name = strip_sql_prefix(base_name).upper()[:31]

            open_fn = gzip.open if str(csv_file).endswith(".gz") else open

            try:
                with open_fn(csv_file, "rt", encoding="utf-8") as f:
                    df = pd.read_csv(f)
            except Exception:
                logger.warning("REPORT excel: failed to read %s, skip", csv_file.name)
                continue

            sheet_frames.setdefault(sheet_name, []).append(df)

        summary_rows = []

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

            for sheet_name, frames in sheet_frames.items():
                df = pd.concat(frames, ignore_index=True)
                row_count = len(df)

                if row_count > 1_048_576:
                    logger.warning("REPORT excel: row limit exceeded, skip | %s rows=%d", sheet_name, row_count)
                    continue

                df.to_excel(writer, sheet_name=sheet_name, index=False)
                summary_rows.append({"sheet_name": sheet_name, "rows": row_count})

                ws = writer.book[sheet_name]
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

                # 헤더 스타일
                header_fill = PatternFill("solid", fgColor="D9E1F2")
                header_font = Font(bold=True)
                for col_idx in range(1, len(df.columns) + 1):
                    c = ws.cell(row=1, column=col_idx)
                    c.fill = header_fill
                    c.font = header_font

                # 컬럼 너비
                for col_idx, col_name in enumerate(df.columns, start=1):
                    col_series = df.iloc[:, col_idx - 1].astype(str)
                    str_max = col_series.str.len().max()
                    max_len = max(str_max if pd.notna(str_max) else 0, len(str(col_name)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(int(max_len * 1.2) + 2, 50)

                if len(frames) > 1:
                    logger.info("REPORT excel sheet: %s (%d rows, merged from %d files)", sheet_name, row_count, len(frames))
                else:
                    logger.info("REPORT excel sheet: %s (%d rows)", sheet_name, row_count)

            # SUMMARY 시트
            if summary_rows:
                import pandas as pd
                summary_df = pd.DataFrame(summary_rows)
                summary_df.insert(0, "no", range(1, len(summary_df) + 1))
                summary_df.to_excel(writer, sheet_name="SUMMARY", index=False)

                ws_sum = writer.book["SUMMARY"]
                ws_sum.freeze_panes = "A2"
                ws_sum.auto_filter.ref = ws_sum.dimensions

                h_fill = PatternFill("solid", fgColor="BDD7EE")
                h_font = Font(bold=True)
                for col_idx in range(1, summary_df.shape[1] + 1):
                    c = ws_sum.cell(row=1, column=col_idx)
                    c.fill = h_fill
                    c.font = h_font

                for col_idx, col_name in enumerate(summary_df.columns, start=1):
                    values = [len(str(col_name))]
                    values += [len(str(v)) for v in summary_df[col_name] if pd.notna(v)]
                    max_len = max(values)
                    ws_sum.column_dimensions[get_column_letter(col_idx)].width = min(int(max_len * 1.2) + 2, 30)

                existing_sheets = set(writer.book.sheetnames)
                for row_idx in range(2, ws_sum.max_row + 1):
                    sheet_val = ws_sum.cell(row=row_idx, column=2).value
                    if sheet_val and sheet_val in existing_sheets:
                        ws_sum.cell(row=row_idx, column=2).hyperlink = f"#'{sheet_val}'!A1"
                        ws_sum.cell(row=row_idx, column=2).style = "Hyperlink"

                wb = writer.book
                wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(ws_sum)))

        logger.info("REPORT excel done | %s (%.2fs)", output_path.name, time.time() - start)

    except ImportError as e:
        logger.error("REPORT excel: package not installed (%s) -> pip install pandas openpyxl", e)
        raise
    except Exception as e:
        logger.exception("REPORT excel generation failed: %s", e)
        raise
